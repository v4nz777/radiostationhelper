import code
import os
from openpyxl import load_workbook

from datetime import datetime
from .models import Invoice, Ad, Customer, InvoiceTransmittal
from mainapp.models import User
from django.core.exceptions import ObjectDoesNotExist

from .environments.love_malaybalay import invoice, invoice_append, collection_append_add



def create_invoice(contract, invoice_no, invoice_from, invoice_till):
        
    products = Ad.objects.filter(code=contract)
    
    advertiser = Ad.objects.get(code=contract).customer
    account_executive = Ad.objects.get(code=contract).account_executive


    #create new instance
    Invoice.objects.create(
        invoice_no=invoice_no,
        from_contract=contract,
        advertiser=advertiser,
        account_executive=account_executive,
        invoice_date=datetime.now(),
        applicable_month_from=datetime.strptime(invoice_from, "%Y-%m-%d"),
        applicable_month_to=datetime.strptime(invoice_till, "%Y-%m-%d"),
        )
    #set products to many to many of instance
    created_invoice = Invoice.objects.get(invoice_no=invoice_no)
    for product in products:
        created_invoice.for_ads.add(product)
        created_invoice.save()
    
    template = os.path.abspath('office/excel/invoice/template-invoice.xlsx')

    workbook = load_workbook(template)

    sheet = workbook['subject']

    _invoice = Invoice.objects.get(invoice_no=invoice_no)
    
    #TITLE
    company = _invoice.advertiser.company.title()

    #ADDRESS
    address = _invoice.advertiser.address

    #DATE
    date = datetime.now().strftime('%d/%m/%Y')

    #PARTICULAR
    app_date_from = _invoice.applicable_month_from
    app_date_to = _invoice.applicable_month_to

    if app_date_from.strftime('%Y') == app_date_to.strftime('%Y'):
        particular = app_date_from.strftime('%B %d') + '-' + app_date_to.strftime('%B %d, %Y')
    else:
        particular = app_date_from.strftime('%B %d, %Y') + '-' + app_date_to.strftime('%B %d, %Y')

    #AMOUNT
    amount = Ad.objects.get(code=_invoice.from_contract).amount

    #IN-CHARGE
    #
    in_charge = 'Florence Maynopas-Dumalag'

    invoice(sheet, company, address, date, particular, amount, in_charge)

    workbook.save(f'office/excel/invoice/{advertiser.company}{invoice_from}{invoice_till}.xlsx')

    invoice_transmittal_append(_invoice)


 
def invoice_transmittal_append(_invoice):
    
    month = _invoice.invoice_date.strftime('%b')
    year = _invoice.invoice_date.strftime('%Y')
    
    
    try:#do nothing if already exist
        if InvoiceTransmittal.objects.get(month=month, year=year):  
            pass

    except ObjectDoesNotExist: #save object if does not exist
        InvoiceTransmittal.objects.create(month=month, year=year)
    
    ###Add Invoice to M2M List
    _open = InvoiceTransmittal.objects.get(month=month, year=year)
    _open.invoices.add(_invoice)
    _open.save()



    #then add to excel
    if not os.path.exists(f'office/excel/invoice-transmittal/{month} {year}.xlsx'):
        workbook = load_workbook('office/excel/invoice-transmittal/billing-invoice-transmittal.xlsx')
        print('WB DOESNT EXIST, TEMPLATE LOADED...')
    else:
        workbook = load_workbook(f'office/excel/invoice-transmittal/{month} {year}.xlsx')
        print('WB EXISTS, LOAD WB...')

    sheet = workbook["subject"] 

    monthyear = month + " " + year
    invoice_date = _invoice.invoice_date
    invoice_no = _invoice.invoice_no
    advertiser = _invoice.advertiser.company.title()
    particular = _invoice.applicable_month_from.strftime('%b-%d-%Y') + " to " + _invoice.applicable_month_to.strftime('%b-%d-%Y')
    amount = Ad.objects.get(code=_invoice.from_contract).amount
    
    invoice_append(sheet, monthyear, invoice_date, invoice_no, advertiser, particular, amount)
    workbook.save(f'office/excel/invoice-transmittal/{month} {year}.xlsx')


def collection_append(invoice_no):
    _invoice = Invoice.objects.get(invoice_no=invoice_no)
    month = _invoice.invoice_date.strftime('%b')
    year = _invoice.invoice_date.strftime('%Y')

    monthyear = month + " " + year
    invoice_date = _invoice.invoice_date.strftime('%m/%d/%Y')
    amount = _invoice.received
    advertiser = _invoice.advertiser.company.title()
    try: 
        ae = f'{_invoice.account_executive.first_name.title()} {_invoice.account_executive.last_name.title()}'
    except AttributeError:#if office sales
        ae = 'OS'
    or_date = _invoice.or_date.strftime('%m/%d/%Y')
    or_number = _invoice.or_number
    applicable_month = _invoice.applicable_month_from.strftime('%b %d,%Y') + "-" + _invoice.applicable_month_to.strftime('%b %d,%Y')
    remarks = 'PAID'
    if not _invoice.paid:
        remarks = 'UNPAID'

    if not os.path.exists(f'office/excel/collection-reports/{month} {year} - Collection Report.xlsx'):
        workbook = load_workbook('office/excel/collection-reports/collection-report-template.xlsx')
        
    else:
        workbook = load_workbook(f'office/excel/collection-reports/{month} {year} - Collection Report.xlsx')
        
    sheet = workbook["subject"]
    

    collection_append_add(sheet,monthyear,invoice_date,int(invoice_no),amount,advertiser,ae,or_date,or_number,applicable_month,remarks)


    workbook.save(f'office/excel/collection-reports/{month} {year} - Collection Report.xlsx')
    

