from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def invoice(sheet, advertiser, address, date, particular, amount, in_charge):
    #TITLE
    sheet['A4'] = advertiser

    #ADDRESS
    sheet['A5'] = address

    #DATE
    sheet['I4'] = date

    #PARTICULAR
    sheet['A11'] = particular

    #AMOUNT
    sheet['F12'] = amount
    sheet['I12'] = amount

    #IN-CHARGE
    sheet['A24'] = in_charge


def invoice_append(sheet, monthyear, invoice_date, invoice_no, advertiser, particular, amount):
    print(sheet.max_row)
    #MONTHYEAR
    sheet["B4"] = monthyear

    sheet.append([invoice_date, invoice_no, advertiser, particular, amount])
    sheet[f"A{sheet.max_row}"].alignment = Alignment(horizontal='left')
    sheet[f"B{sheet.max_row}"].alignment = Alignment(horizontal='center')
    sheet[f"E{sheet.max_row}"].number_format = '₱#,##0.00'
    sheet[f"E{sheet.max_row}"].alignment = Alignment(horizontal='left')

def collection_append_add(sheet,monthyear,invoice_date,invoice_no,amount,advertiser,ae,or_date,or_number,applicable_month,remarks):
    sheet["B3"] = monthyear

    sheet.append([invoice_date,invoice_no,amount,advertiser,ae,or_date,or_number,applicable_month,remarks])
    sheet[f"F{sheet.max_row}"].alignment = Alignment(horizontal='center')
    sheet[f"G{sheet.max_row}"].alignment = Alignment(horizontal='center')
    sheet[f"I{sheet.max_row}"].alignment = Alignment(horizontal='center')



 